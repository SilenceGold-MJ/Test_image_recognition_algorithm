# -*- coding: utf-8 -*-
# !/usr/bin/python3
import openpyxl, os, datetime
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.styles import PatternFill  # 导入填充模块
from framework.logger import Logger

logger = Logger(logger="ExportExcle").getlog()


def ExportExcle(addr, dic, title, n):
    # n是判断结果颜色的列
    # 设置文件 mingc
    # 打开文件
    wb = openpyxl.Workbook()  # load_workbook(addr)
    # 创建一张新表
    # ws = wb.create_sheet()
    ws = wb.active
    ws.title = title
    # 第一行输入
    sheet_names = wb.sheetnames
    sheet1 = wb[title]  # 打开第一个 sheet 工作表
    Excle_head = list(dic.keys())
    column = 1
    for i in Excle_head:
        sheet1.cell(row=1, column=column, value=i).font = Font(bold=True)  # 实际检索结果
        column = column + 1
    # ws.append(csv_head)
    list_value = list(dic.values())
    for i in range(len(list_value)):
        if i == n:
            # ft = Font(color=list_value[i][1][0])
            # sheet1.cell(row=list_value[0] + 1, column=i + 1, value=list_value[i][0]).font = ft


            fille = PatternFill('solid', fgColor=list_value[i][1][0])  # 设置填充颜色为 橙色
            font = Font(u'宋体', size=11, bold=False, italic=False, strike=False, color=list_value[i][1][1])  # 设置字体样式
            sheet1.cell(row=list_value[0] + 1, column=i + 1, value="").fill = fille  # 序列
            sheet1.cell(row=list_value[0] + 1, column=i + 1, value=list_value[i][0]).font = font  # 序列


        else:
            sheet1.cell(row=list_value[0] + 1, column=i + 1, value=list_value[i])  # 序列
    wb.save(addr)


def SummaryExcle(addr, dic, title, n):  # excle文件路径、输入数据，写入工作表名称，结果颜色列
    # n是判断结果颜色的列
    wb = load_workbook(addr)  # load_workbook(addr)# 打开文件
    # ws = wb.create_sheet()# 创建一张新表
    # ws = wb.active
    # ws.title = '汇总页'
    sheet_names = wb.sheetnames  # 获取所有sheet
    if title in sheet_names:
        sheet1 = wb[title]  # 打开测试记录 sheet 工作表
    else:
        wb.create_sheet(title, 0)  # 插入到最开始的位置
        sheet1 = wb[title]
        Excle_head = list(dic.keys())
        column = 1
        for i in Excle_head:
            sheet1.cell(row=1, column=column, value=i).font = Font(bold=True)  # 实际检索结果
            column = column + 1
    list_value = list(dic.values())

    for i in range(len(list_value)):
        if i == n:
            # ft = Font(color=list_value[i][1][0])
            # sheet1.cell(row=list_value[0] + 1, column=i + 1, value=list_value[i][0]).font = ft


            fille = PatternFill('solid', fgColor=list_value[i][1][0])  # 设置填充颜色为 橙色
            font = Font(u'宋体', size=11, bold=False, italic=False, strike=False, color=list_value[i][1][1])  # 设置字体样式
            sheet1.cell(row=list_value[0] + 1, column=i + 1, value="").fill = fille  # 序列
            sheet1.cell(row=list_value[0] + 1, column=i + 1, value=list_value[i][0]).font = font  # 序列


        else:
            sheet1.cell(row=list_value[0] + 1, column=i + 1, value=list_value[i])  # 序列
    wb.save(addr)


def VerticalExcle(addr, dic, title, n):  # excle文件路径、输入数据，写入工作表名称，结果颜色列
    # n是判断结果颜色的列
    wb = load_workbook(addr)  # load_workbook(addr)# 打开文件
    # ws = wb.create_sheet()# 创建一张新表
    # ws = wb.active
    # ws.title = '汇总页'
    sheet_names = wb.sheetnames  # 获取所有sheet
    sheet1 = wb[title]  # 打开测试记录 sheet 工作表
    list_value = list(dic.values())
    sheet1.cell(1, column=n + 1, value='数据汇总').font = Font(bold=True)
    for i in range(len(list_value)):
        sheet1.cell(row=i + 2, column=n + 1, value=list_value[i])  # 序列
    sheet1.cell(13, column=n + 1, value='于' + datetime.datetime.now().strftime('%Y{y}%m{m}%d{d}').format(y='年', m='月',
                                                                                                         d='日')).font = Font(
        bold=True)
    wb.save(addr)


def getnum(filepath, title):
    Excelpath = os.getcwd() + "/" + filepath
    if os.path.exists(Excelpath):
        wb = load_workbook(filepath)
        sheet_names = wb.sheetnames  # 得到工作簿的所有工作表名 结果： ['Sheet1', 'Sheet2', 'Sheet3']
        sheet1 = wb[title]  # 打开第一个 sheet 工作表
        list_A = []
        for a in sheet1["A"]:
            list_A.append(a.value)
        if None in list_A:
            return list_A.index(None)
        else:
            return len(list_A)
    else:
        return 1
